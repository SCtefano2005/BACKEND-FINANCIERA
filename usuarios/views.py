from rest_framework import viewsets, status, permissions
from rest_framework.response import Response
from rest_framework.views import APIView
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from .models import Cliente, Proveedor, FacturaCliente, FacturaProveedor
from .serializers import (
    ClienteSerializer, ProveedorSerializer, FacturaClienteSerializer, FacturaProveedorSerializer, UsuarioReadSerializer, CustomTokenObtainPairSerializer
)
from django.template.loader import render_to_string
from xhtml2pdf import pisa  # Cambiado para generación de PDF
import openpyxl
from openpyxl.styles import Font, Alignment
from asgiref.sync import async_to_sync
from channels.layers import get_channel_layer
from django.contrib.auth import get_user_model
from rest_framework_simplejwt.views import TokenObtainPairView

Usuario = get_user_model()

class CustomTokenObtainPairView(TokenObtainPairView):
    serializer_class = CustomTokenObtainPairSerializer
    
class UsuarioViewSet(viewsets.ModelViewSet):
    queryset = Usuario.objects.all()
    serializer_class = UsuarioReadSerializer

    def get_permissions(self):
        if self.request.user.rol == "ADMIN":
            # Administradores tienen CRUD completo para usuarios
            return [permissions.IsAuthenticated()]
        return [permissions.IsAdminUser()]


class ClienteViewSet(viewsets.ModelViewSet):
    queryset = Cliente.objects.all()
    serializer_class = ClienteSerializer

    def get_permissions(self):
        if self.request.user.rol == "GERENTE":
            if self.action in ["list", "retrieve"]:
                return [permissions.IsAuthenticated()]
            return [permissions.IsAdminUser()]
        elif self.request.user.rol == "ADMIN":
            return [permissions.IsAuthenticated()]
        return [permissions.IsAdminUser()]


class ProveedorViewSet(viewsets.ModelViewSet):
    queryset = Proveedor.objects.all()
    serializer_class = ProveedorSerializer

    def get_permissions(self):
        if self.request.user.rol == "GERENTE":
            if self.action in ["list", "retrieve"]:
                return [permissions.IsAuthenticated()]
            return [permissions.IsAdminUser()]
        elif self.request.user.rol == "ADMIN":
            return [permissions.IsAuthenticated()]
        return [permissions.IsAdminUser()]


class FacturaClienteViewSet(viewsets.ModelViewSet):
    queryset = FacturaCliente.objects.all()
    serializer_class = FacturaClienteSerializer

    def get_permissions(self):
        if self.request.user.rol == "CONTADOR":
            return [permissions.IsAuthenticated()]
        elif self.request.user.rol == "GERENTE":
            if self.action in ["list", "retrieve"]:
                return [permissions.IsAuthenticated()]
            return [permissions.IsAdminUser()]
        elif self.request.user.rol == "ADMIN":
            return [permissions.IsAuthenticated()]
        return [permissions.IsAdminUser()]


class FacturaProveedorViewSet(viewsets.ModelViewSet):
    queryset = FacturaProveedor.objects.all()
    serializer_class = FacturaProveedorSerializer

    def get_permissions(self):
        if self.request.user.rol == "CONTADOR":
            return [permissions.IsAuthenticated()]
        elif self.request.user.rol == "GERENTE":
            if self.action in ["list", "retrieve"]:
                return [permissions.IsAuthenticated()]
            return [permissions.IsAdminUser()]
        elif self.request.user.rol == "ADMIN":
            return [permissions.IsAuthenticated()]
        return [permissions.IsAdminUser()]


class FacturaPDFView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request, factura_id):
        if request.user.rol not in ["GERENTE", "ADMIN"]:
            return Response({'error': 'No tiene permisos para realizar esta acción.'}, status=status.HTTP_403_FORBIDDEN)
        
        factura = get_object_or_404(FacturaCliente, id=factura_id)
        html_string = render_to_string('factura_cliente.html', {'factura': factura})
        
        # Genera el PDF usando xhtml2pdf
        pdf_file = HttpResponse(content_type='application/pdf')
        pdf_file['Content-Disposition'] = f'attachment; filename="Factura_{factura.numero_factura}.pdf"'
        pisa_status = pisa.CreatePDF(html_string, dest=pdf_file)
        
        if pisa_status.err:
            return Response({'error': 'Error al generar el PDF'}, status=status.HTTP_500_INTERNAL_SERVER_ERROR)
        
        return pdf_file


class GenerarExcelFacturasView(APIView):
    permission_classes = [permissions.IsAuthenticated]

    def get(self, request):
        if request.user.rol != "CONTADOR":
            return Response({'error': 'No tiene permisos para realizar esta acción.'}, status=status.HTTP_403_FORBIDDEN)

        facturas = FacturaCliente.objects.all()

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Facturas"

        headers = ["ID", "Número de Factura", "Cliente", "Monto Total", "Estado", "Fecha de Emisión"]
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num)
            cell.value = header
            cell.font = Font(bold=True)
            cell.alignment = Alignment(horizontal="center", vertical="center")

        row = 2
        for factura in facturas:
            ws.cell(row=row, column=1).value = factura.id
            ws.cell(row=row, column=2).value = factura.numero_factura
            ws.cell(row=row, column=3).value = factura.cliente.nombre
            ws.cell(row=row, column=4).value = factura.monto_total
            ws.cell(row=row, column=5).value = factura.estado
            ws.cell(row=row, column=6).value = factura.fecha_emision.strftime('%Y-%m-%d')
            row += 1

        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = 'attachment; filename="facturas.xlsx"'
        wb.save(response)
        return response


def notificar_factura_vencimiento(factura):
    channel_layer = get_channel_layer()
    mensaje = f"La factura {factura.numero_factura} de {factura.cliente.nombre} está próxima a vencer el {factura.fecha_vencimiento.strftime('%Y-%m-%d')}."
    async_to_sync(channel_layer.group_send)(
        "notificaciones_facturas",
        {
            "type": "enviar_notificacion",
            "mensaje": mensaje,
        },
    )
